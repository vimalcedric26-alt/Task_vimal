import openpyxl


class ExcelData:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def load_excel_wrkbook(self):
        workbook = openpyxl.load_workbook(self.filename)
        worksheet = workbook[self.sheetname]
        return workbook, worksheet

    def get_data_(self, worksheet):
        rows = worksheet.max_row
        columns = worksheet.max_column

        all_data = []

        for each_row in range(2, rows + 1):  # skip header
            row_data = []
            for each_col in range(1, columns + 1):
                cell_value = worksheet.cell(row=each_row, column=each_col).value
                row_data.append(cell_value)

            all_data.append((each_row, row_data))

        return all_data

    def update_result(self, worksheet, row_number, result):
        worksheet.cell(row=row_number, column=4).value = result

    def save_workbook(self, workbook):
        workbook.save(self.filename)